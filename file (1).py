#!/usr/bin/env python3
"""
compare_lifts.py — All‑Python CLI for lift comparison and Excel generation.

Features
- Load consultant + vendor offers from plain text files or a prebuilt JSON payload.
- Light heuristics to parse key: value lines into structured sections.
- Interactive review/edit in terminal (optional).
- Generate formatted Excel (COMPARISON_SHEET_OUTPUT.xlsx) with merges, fills, borders, totals,
  payment terms, delivery program, remarks, and approval matrix.

Quick start
  pip install -r requirements.txt
  python compare_lifts.py wizard          # guided flow, paste or load files
  python compare_lifts.py from-json comparison_data.json  # noninteractive
  python compare_lifts.py excel comparison_data.json      # directly write Excel

Schema (internal)
  {
    "projectInfo": { tel, fax, pobox, project, material, prNo, date },
    "sections": {
      "<sectionKey>": {
        "<FIELD>": { "consultant": str, "KONE": str, "TKE": str, "EEE": str, "AG MELCO": str },
        ...
      },
      ...
    },
    "pricing": { "<PRICE_ITEM>": { "KONE": str|num, "TKE": str|num, "EEE": str|num, "AG MELCO": str|num } },
    "paymentTerms": { "KONE": [str], "TKE": [str], "EEE": [str], "AG MELCO": [str] },
    "delivery":    { "KONE": [str], "TKE": [str], "EEE": [str], "AG MELCO": [str] }
  }
"""

import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import typer
from rich import box
from rich.console import Console
from rich.prompt import Confirm, Prompt
from rich.table import Table
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = typer.Typer(add_completion=False)
console = Console()

# ---------------------------------------------------------------------------
# Domain Constants
# ---------------------------------------------------------------------------

VENDORS = ["KONE", "TKE", "EEE", "AG MELCO"]

SPEC_FIELDS = [
    "CAPACITY", "SPEED", "DOOR TYPE", "DOOR SIZE (W X H)", "SHAFT SIZE (W X D)",
    "CABIN SIZE (W X D X H)", "OVER HEAD HEIGHT", "PIT DEPTH", "NO. OF LIFTS",
    "Lift code", "Machine location", "Operation", "No. of Stops", "Travel Height",
    "Car wall", "Front Wall", "Ceiling", "Mirror", "Hand rail", "Skirting",
    "Decoration", "Door Material", "Sill Material", "COP Panel", "LOP",
    "Landing Jamb In Ground Floor", "Landing Jamb In Other Floors", "Hall Indicator", "Made"
]

FIREMAN_FIELDS = [
    "CAPACITY", "SPEED", "DOOR TYPE", "DOOR SIZE (W X H)", "SHAFT SIZE (W X D)",
    "CABIN SIZE (W X D X H)", "OVER HEAD HEIGHT", "PIT DEPTH", "NO. OF LIFTS",
    "Lift code", "Machine location", "Operation", "No. of Stops", "Travel Height",
    "Car wall", "Front Wall", "Ceiling", "Mirror", "Hand rail", "Skirting",
    "Decoration", "Door Material", "Sill Material", "COP Panel", "LOP",
    "Landing Jamb In Ground Floor", "Landing Jamb In Other Floors",
    "Hall Indicator (Ground Floor)", "Hall Indicator (Other Floor)", "Made"
]

SECTIONS = [
    {"key": "towerA_pl12", "label": "TOWER A - PL1 , PL2 ", "type": "passenger"},
    {"key": "towerA_pl345", "label": "TOWER A - PL3 , PL4, PL5", "type": "passenger"},
    {"key": "towerA_fl", "label": "TOWER A - FIREMAN LIFT ", "type": "fireman"},
    {"key": "towerB_pl12", "label": "TOWER B - PL1 , PL2", "type": "passenger"},
    {"key": "towerB_pl345", "label": "TOWER B - PL3 , PL4, PL5", "type": "passenger"},
    {"key": "towerB_fl", "label": "TOWER B - FIREMAN LIFT ", "type": "fireman"},
    {"key": "towerC_pl12", "label": "TOWER C - PL1 , PL2", "type": "passenger"},
    {"key": "towerC_pl3", "label": "TOWER C - PL3", "type": "passenger"},
    {"key": "towerC_fl", "label": "TOWER C - FIREMAN LIFT ", "type": "fireman"},
]

PRICE_ITEMS = [
    "Tower A Passenger Lift (PL1, PL2)", "Tower A Passenger Lift (PL3, PL4, PL5)", "Tower A Firemen Lift (FL1)",
    "Tower B Passenger Lift (PL1, PL2)", "Tower B Passenger Lift (PL3, PL4, PL5)", "Tower B Firemen Lift (FL1)",
    "Tower C Passenger Lift (PL1, PL2)", "Tower C Passenger Lift (PL3)", "Tower C Firemen Lift (FL1)",
    "Monitoring System", "Separator Beam", "Multimedia", "IP ratings"
]

# Mapping for Excel label variants
PRICE_ITEMS_DISPLAY = {
    "Tower A Passenger Lift (PL1, PL2)": "Tower A Passenger Lift (PL1 , PL2)",
    "Tower A Passenger Lift (PL3, PL4, PL5)": "Tower A Passenger Lift (PL3 , PL4, PL5)",
    "Tower A Firemen Lift (FL1)": "Tower A Firemen Lift (FL1)",
    "Tower B Passenger Lift (PL1, PL2)": "Tower B Passenger Lift (PL1 , PL2)",
    "Tower B Passenger Lift (PL3, PL4, PL5)": "Tower B Passenger Lift (PL3 , PL4, PL5)",
    "Tower B Firemen Lift (FL1)": "Tower B Firemen Lift (FL1)",
    "Tower C Passenger Lift (PL1, PL2)": "Tower C Passenger Lift (PL1 , PL2)",
    "Tower C Passenger Lift (PL3)": "Tower C Passenger Lift (PL3)",
    "Tower C Firemen Lift (FL1)": "Tower C Firemen Lift (FL1)",
    "Monitoring System": "Monitoring System",
    "Separator Beam": "Seperator Beam",
    "Multimedia": "Multimedia",
    "IP ratings": "IP ratings",
}

# ---------------------------------------------------------------------------
# Utilities and Models
# ---------------------------------------------------------------------------

def today_ddmmyyyy() -> str:
    from datetime import date
    d = date.today()
    return f"{d.day:02d}/{d.month:02d}/{d.year}"

def initial_sections() -> Dict[str, Dict[str, Dict[str, str]]]:
    data = {}
    for s in SECTIONS:
        fields = FIREMAN_FIELDS if s["type"] == "fireman" else SPEC_FIELDS
        data[s["key"]] = {}
        for f in fields:
            data[s["key"]][f] = { "consultant": "", "KONE": "", "TKE": "", "EEE": "", "AG MELCO": "" }
    return data

def initial_pricing() -> Dict[str, Dict[str, str]]:
    p = {}
    for item in PRICE_ITEMS:
        p[item] = { v: "" for v in VENDORS }
    return p

def default_payload() -> Dict:
    return {
        "projectInfo": {
            "tel": "02 883 6543",
            "fax": "02 6263460",
            "pobox": "45195",
            "project": "RADIANT BRIDGES TOWERS",
            "material": "ELEVATOR",
            "prNo": "",
            "date": today_ddmmyyyy(),
        },
        "sections": initial_sections(),
        "pricing": initial_pricing(),
        "paymentTerms": { v: [] for v in VENDORS },
        "delivery": { v: [] for v in VENDORS },
    }

SPEC_KEY_NORMALISE = {
    # map common variants to canonical keys
    "overhead height": "OVER HEAD HEIGHT",
    "over head height": "OVER HEAD HEIGHT",
    "pit": "PIT DEPTH",
    "no of lifts": "NO. OF LIFTS",
    "no. of lifts": "NO. OF LIFTS",
    "door size": "DOOR SIZE (W X H)",
    "shaft size": "SHAFT SIZE (W X D)",
    "cabin size": "CABIN SIZE (W X D X H)",
    "hall indicator (ground floor)": "Hall Indicator (Ground Floor)",
    "hall indicator (other floor)": "Hall Indicator (Other Floor)",
}

def normalise_field_name(name: str) -> str:
    n = re.sub(r"\s+", " ", name).strip()
    key = n.lower()
    return SPEC_KEY_NORMALISE.get(key, n)

KV_LINE = re.compile(r"^\s*([A-Za-z0-9()\/\-\.\s]+?)\s*[:\-]\s*(.+?)\s*$")

def parse_text_block_to_kv(text: str) -> Dict[str, str]:
    """
    Heuristic parser: lines like 'CAPACITY: 1000 kg' -> { 'CAPACITY': '1000 kg' }
    """
    out: Dict[str, str] = {}
    for raw in text.splitlines():
        m = KV_LINE.match(raw)
        if not m:
            continue
        k, v = m.group(1), m.group(2)
        k = normalise_field_name(k)
        out[k] = v.strip()
    return out

def apply_kv_to_sections(sections: Dict, sec_key: str, role: str, kv: Dict[str, str]):
    if sec_key not in sections:
        return
    for field in sections[sec_key].keys():
        # exact match
        if field in kv:
            sections[sec_key][field][role] = kv[field]
            continue
        # relaxed match ignoring case/spaces
        f_norm = re.sub(r"\s+", "", field).lower()
        found = None
        for k, v in kv.items():
            if re.sub(r"\s+", "", k).lower() == f_norm:
                found = v
                break
        if found is not None:
            sections[sec_key][field][role] = found

def parse_money(s: str) -> float:
    if s is None:
        return 0.0
    st = str(s).strip()
    if st.startswith("(") and st.endswith(")"):
        st = "-" + st[1:-1]
    st = st.replace(",", "")
    st = re.sub(r"[^0-9.\-]", "", st)
    try:
        return float(st)
    except ValueError:
        return 0.0

# ---------------------------------------------------------------------------
# Excel Writer
# ---------------------------------------------------------------------------

def make_excel(payload: Dict, out_file: str = "COMPARISON_SHEET_OUTPUT.xlsx") -> str:
    pi = payload.get("projectInfo", {})
    sections = payload.get("sections", {})
    pricing = payload.get("pricing", {})
    pay_terms = payload.get("paymentTerms", {})
    delivery = payload.get("delivery", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    def fill(hex_): return PatternFill("solid", fgColor=hex_)
    def font_(bold=False, size=10, color="000000"): return Font(bold=bold, size=size, color=color, name="Calibri")
    def align_(h="left", v="center", wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    HEADER_FILL = fill("FFD6DCE4")
    SUBHDR_FILL = fill("FFBFBFBF")
    ALT_FILL = fill("FFEEEEEE")

    def set_cell(row, col, value, bold=False, fill_=None, h_align="left", wrap=True, size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font_(bold=bold, size=size)
        c.alignment = align_(h=h_align, wrap=wrap)
        if fill_: c.fill = fill_
        c.border = border_all
        return c

    def merge_set(row, c1, c2, value, bold=False, fill_=None, h_align="center", size=10):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border_all
            if fill_: cell.fill = fill_
        main_cell = ws.cell(row=row, column=c1, value=value)
        main_cell.font = font_(bold=bold, size=size)
        main_cell.alignment = align_(h=h_align)

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 53
    ws.column_dimensions["D"].width = 53
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 53
    ws.column_dimensions["G"].width = 50

    row = 1
    set_cell(row, 2, f"Tel.: {pi.get('tel','')}")
    row = 2
    set_cell(row, 2, f"Fax.: {pi.get('fax','')}")
    set_cell(row, 7, pi.get("date",""), h_align="center")
    row = 3
    set_cell(row, 2, f"P.O.Box: {pi.get('pobox','')}")
    row = 4
    merge_set(row, 2, 7, "COMPARISON SHEET - LIFT", bold=True, size=14)
    row = 5
    set_cell(row, 2, f"PROJECT: {pi.get('project','')}", bold=True)
    row = 6
    set_cell(row, 2, f"MATERIAL/WORK: {pi.get('material','')}", bold=True)
    row = 7
    set_cell(row, 2, f"PR NO.: {pi.get('prNo','')}", bold=True)

    # Sections
    def write_section(start_row, sec_key, sec_label, sec_type):
        r = start_row
        merge_set(r, 2, 7, sec_label, bold=True, fill_=HEADER_FILL, h_align="left")
        r += 1
        for ci, label in enumerate(["Specification","Consultant ","KONE ","TKE","EEE","AG MELCO"], start=2):
            set_cell(r, ci, label, bold=True, fill_=SUBHDR_FILL, h_align="center")
        r += 1
        fields = FIREMAN_FIELDS if sec_type == "fireman" else SPEC_FIELDS
        sec_data = sections.get(sec_key, {})
        for i, field in enumerate(fields):
            row_fill = ALT_FILL if i % 2 == 0 else None
            set_cell(r, 2, field, fill_=row_fill)
            vals = sec_data.get(field, {})
            set_cell(r, 3, vals.get("consultant",""), fill_=row_fill)
            set_cell(r, 4, vals.get("KONE",""), fill_=row_fill)
            set_cell(r, 5, vals.get("TKE",""), fill_=row_fill)
            set_cell(r, 6, vals.get("EEE",""), fill_=row_fill)
            set_cell(r, 7, vals.get("AG MELCO",""), fill_=row_fill)
            r += 1
        return r

    current = 8
    for s in SECTIONS:
        current = write_section(current, s["key"], s["label"], s["type"])

    # Spacer
    merge_set(current, 2, 7, "")
    current += 1

    # Commercial header
    for ci, label in enumerate(["S. No.","DESCRIPTION","KONE","TKE","EEE","AG MELCO"], start=2):
        set_cell(current, ci, label, bold=True, fill_=SUBHDR_FILL, h_align="center")
    current += 1

    # Prices
    price_row_idx: Dict[str, int] = {}
    for i, item in enumerate(PRICE_ITEMS, start=1):
        prow = pricing.get(item, {})
        vals = [prow.get(v, "") for v in VENDORS]

        cleaned: List[Optional[float]] = []
        for v in vals:
            st = str(v).strip()
            if st.startswith("(") and st.endswith(")"):
                st = "-" + st[1:-1]
            st = st.replace(",", "")
            st2 = re.sub(r"[^0-9.\-]", "", st)
            try:
                num = float(st2) if st2 not in ("", "-", ".", "-.", ".-") else None
            except ValueError:
                num = None
            cleaned.append(num if num is not None else v)

        set_cell(current, 2, i, h_align="center")
        set_cell(current, 3, PRICE_ITEMS_DISPLAY.get(item, item))
        for ci, v in enumerate(cleaned):
            c = set_cell(current, 4 + ci, v, h_align="center")
            if isinstance(v, (int, float, float)):
                ws.cell(row=current, column=4 + ci).number_format = "#,##0"
        price_row_idx[item] = current
        current += 1

    # Total for first nine items (lifts)
    set_cell(current, 2, "TOTAL", bold=True)
    set_cell(current, 3, "", bold=True)
    if PRICE_ITEMS[:9]:
        pr_start = price_row_idx[PRICE_ITEMS[0]]
        pr_end = price_row_idx[PRICE_ITEMS[8]]
        for ci, col_letter in enumerate(["D","E","F","G"]):
            c = ws.cell(row=current, column=4+ci, value=f"=SUM({col_letter}{pr_start}:{col_letter}{pr_end})")
            c.font = font_(bold=True)
            c.alignment = align_(h="center")
            c.border = border_all
            c.number_format = "#,##0"
    current += 2

    # Payment terms
    merge_set(current, 2, 7, "PAYMENT TERMS", bold=True, fill_=HEADER_FILL, h_align="left")
    current += 1
    set_cell(current, 2, "S. No.", bold=True, fill_=SUBHDR_FILL, h_align="center")
    set_cell(current, 3, "KONE", bold=True, fill_=SUBHDR_FILL, h_align="center")
    ws.merge_cells(start_row=current, start_column=3, end_row=current, end_column=4)
    set_cell(current, 5, "TKE", bold=True, fill_=SUBHDR_FILL, h_align="center")
    set_cell(current, 6, "EEE", bold=True, fill_=SUBHDR_FILL, h_align="center")
    set_cell(current, 7, "AG MELCO", bold=True, fill_=SUBHDR_FILL, h_align="center")
    current += 1

    max_pay = max((len(pay_terms.get(v, [])) for v in VENDORS), default=0)
    for i in range(max(max_pay, 5)):
        set_cell(current, 2, i + 1, h_align="center")
        kone_terms = pay_terms.get("KONE", [])
        set_cell(current, 3, kone_terms[i] if i < len(kone_terms) else "")
        ws.merge_cells(start_row=current, start_column=3, end_row=current, end_column=4)
        ws.cell(row=current, column=4).border = border_all
        for ci, v in enumerate(["TKE", "EEE", "AG MELCO"]):
            terms = pay_terms.get(v, [])
            set_cell(current, 5 + ci, terms[i] if i < len(terms) else "")
        current += 1

    current += 1
    merge_set(current, 2, 7, "Delivery Program", bold=True, fill_=HEADER_FILL, h_align="left")
    current += 1
    set_cell(current, 2, "S. No.", bold=True, fill_=SUBHDR_FILL, h_align="center")
    set_cell(current, 3, "KONE", bold=True, fill_=SUBHDR_FILL, h_align="center")
    ws.merge_cells(start_row=current, start_column=3, end_row=current, end_column=4)
    set_cell(current, 5, "TKE", bold=True, fill_=SUBHDR_FILL, h_align="center")
    set_cell(current, 6, "EEE", bold=True, fill_=SUBHDR_FILL, h_align="center")
    set_cell(current, 7, "AG MELCO", bold=True, fill_=SUBHDR_FILL, h_align="center")
    current += 1

    max_del = max((len(delivery.get(v, [])) for v in VENDORS), default=0)
    for i in range(max(max_del, 6)):
        set_cell(current, 2, i + 1, h_align="center")
        kone_del = delivery.get("KONE", [])
        set_cell(current, 3, kone_del[i] if i < len(kone_del) else "")
        ws.merge_cells(start_row=current, start_column=3, end_row=current, end_column=4)
        ws.cell(row=current, column=4).border = border_all
        for ci, v in enumerate(["TKE", "EEE", "AG MELCO"]):
            d_list = delivery.get(v, [])
            set_cell(current, 5 + ci, d_list[i] if i < len(d_list) else "")
        current += 1

    current += 1
    merge_set(current, 2, 7, "ENGINEERS RECOMMENDATIONS/REMARKS", bold=True, fill_=HEADER_FILL, h_align="left")
    current += 1
    for _ in range(5):
        merge_set(current, 2, 7, "")
        current += 1

    merge_set(current, 2, 7, "APPROVAL MATRIX", bold=True, fill_=HEADER_FILL, h_align="left")
    current += 2
    approvals = [
        ("PREPARED BY:", "TECHNICAL ENGR."),
        ("CHECKED BY:", "QS ENGR."),
        ("REVIEWED BY:", "MEP MANAGER"),
        ("RECOMMENDED BY:", "PROCUREMENT MGR"),
        ("VERIFIED BY:", "TECHNICAL MGR"),
        ("APPROVED BY:", "PROJECT MANAGER"),
    ]
    for title, role in approvals:
        set_cell(current, 2, title, bold=True)
        merge_set(current, 3, 5, "", h_align="left")
        set_cell(current, 6, role, bold=True)
        current += 2

    ws.freeze_panes = "D10"
    wb.save(out_file)
    return out_file

# ---------------------------------------------------------------------------
# CLI Workflows
# ---------------------------------------------------------------------------

def show_section_table(payload: Dict, section_key: str):
    sec_meta = next(s for s in SECTIONS if s["key"] == section_key)
    sec = payload["sections"][section_key]
    table = Table(title=sec_meta["label"].strip(), show_lines=False, header_style="bold", box=box.SIMPLE)
    table.add_column("Field", style="bold cyan")
    table.add_column("Consultant", style="white")
    for v in VENDORS:
        table.add_column(v, style="white")
    for i, field in enumerate(sec.keys()):
        vals = sec[field]
        row = [field, vals.get("consultant","")] + [vals.get(v,"") for v in VENDORS]
        table.add_row(*row)
    console.print(table)

def interactive_edit(payload: Dict):
    console.rule("[bold]Interactive Review")
    while True:
        # choose section
        section_labels = [f"{i+1}. {s['label'].strip()}" for i, s in enumerate(SECTIONS)]
        console.print("\n".join(section_labels))
        pick = Prompt.ask("Open section number (or 'p' pricing, 'q' to finish)", default="q")
        if pick.lower() == "q":
            break
        if pick.lower() == "p":
            # pricing quick view/edit
            table = Table(title="Pricing Summary", box=box.SIMPLE)
            table.add_column("#")
            table.add_column("Description", style="bold")
            for v in VENDORS:
                table.add_column(v, justify="right")
            for i, item in enumerate(PRICE_ITEMS, start=1):
                row = [str(i), item] + [str(payload["pricing"].get(item, {}).get(v, "")) for v in VENDORS]
                table.add_row(*row)
            console.print(table)
            if Confirm.ask("Edit a price item?", default=False):
                idx = int(Prompt.ask("Item number (1..{})".format(len(PRICE_ITEMS))))
                ven = Prompt.ask("Vendor", choices=VENDORS)
                val = Prompt.ask("New value", default=payload["pricing"][PRICE_ITEMS[idx-1]].get(ven, ""))
                payload["pricing"][PRICE_ITEMS[idx-1]][ven] = val
            continue

        try:
            idx = int(pick) - 1
            sec_meta = SECTIONS[idx]
        except Exception:
            console.print("[red]Invalid choice.[/red]")
            continue
        sec_key = sec_meta["key"]
        show_section_table(payload, sec_key)
        if not Confirm.ask("Edit a field in this section?", default=False):
            continue
        field = Prompt.ask("Type exact field name to edit (case sensitive)")
        if field not in payload["sections"][sec_key]:
            console.print("[red]Field not found in this section.[/red]")
            continue
        who = Prompt.ask("Column to edit", choices=["consultant"] + VENDORS)
        cur = payload["sections"][sec_key][field].get(who, "")
        newv = Prompt.ask("New value", default=cur)
        payload["sections"][sec_key][field][who] = newv

@app.command()
def wizard(
    consultant_file: Optional[Path] = typer.Argument(None, help="Consultant spec text file"),
    kone_file: Optional[Path] = typer.Option(None, help="KONE offer text file"),
    tke_file: Optional[Path] = typer.Option(None, help="TKE offer text file"),
    eee_file: Optional[Path] = typer.Option(None, help="EEE offer text file"),
    agmelco_file: Optional[Path] = typer.Option(None, "--ag", help="AG MELCO offer text file"),
):
    """
    Guided flow:
    - Load default payload
    - Optionally parse text files into sections (heuristic)
    - Interactive review/edit
    - Save JSON and generate Excel
    """
    payload = default_payload()

    # Load text if provided
    texts = {}
    def read_text(path: Optional[Path]) -> str:
        if not path: return ""
        try:
            return path.read_text(encoding="utf-8", errors="ignore")
        except Exception as e:
            console.print(f"[red]Failed to read {path}: {e}[/red]")
            return ""

    texts["consultant"] = read_text(consultant_file)
    texts["KONE"] = read_text(kone_file)
    texts["TKE"] = read_text(tke_file)
    texts["EEE"] = read_text(eee_file)
    texts["AG MELCO"] = read_text(agmelco_file)

    # Optional quick parse (applies same parsed KV to all sections of matching type)
    if any(texts.values()) and Confirm.ask("Attempt heuristic parsing of provided texts?", default=True):
        # Parse consultant spec first, then vendor offers
        con_kv = parse_text_block_to_kv(texts["consultant"])
        for s in SECTIONS:
            apply_kv_to_sections(payload["sections"], s["key"], "consultant", con_kv)
        for v in VENDORS:
            v_kv = parse_text_block_to_kv(texts[v])
            for s in SECTIONS:
                apply_kv_to_sections(payload["sections"], s["key"], v, v_kv)

    # Project info quick edit
    console.rule("[bold]Project Info")
    for k in list(payload["projectInfo"].keys()):
        cur = payload["projectInfo"][k]
        payload["projectInfo"][k] = Prompt.ask(f"{k}", default=str(cur))

    # Interactive data review
    interactive_edit(payload)

   # Payment and delivery entry
console.rule("[bold]Payment / Delivery")
if Confirm.ask("Enter/edit payment terms now?", default=False):
    for v in VENDORS:
        console.print(f"[bold]{v} Payment Terms[/bold]")
        lines: List[str] = []
        console.print("Enter lines (blank line to finish):")
        while True:
            ln = Prompt.ask("", default="")
            if not ln.strip():
                break
            lines.append(ln)
        payload["paymentTerms"][v] = lines

if Confirm.ask("Enter/edit delivery program now?", default=False):
    for v in VENDORS:
        console.print(f"[bold]{v} Delivery Program[/bold]")
        lines: List[str] = []
        console.print("Enter lines (blank line to finish):")
        while True:
            ln = Prompt.ask("", default="")
            if not ln.strip():
                break
            lines.append(ln)
        payload["delivery"][v] = lines

# Save JSON and generate Excel
out_json = Prompt.ask("Save JSON as", default="comparison_data.json")
Path(out_json).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
console.print(f"[green]Saved[/green] {out_json}")

out_xlsx = Prompt.ask("Excel filename", default="COMPARISON_SHEET_OUTPUT.xlsx")
make_excel(payload, out_file=out_xlsx)
console.print(f"[green]Excel written:[/green] {out_xlsx}")
